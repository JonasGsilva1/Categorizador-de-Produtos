"""
Leitura e escrita de arquivos .xlsx.
- Leitura com openpyxl (read_only para eficiência de memória)
- Escrita com xlsxwriter (streaming para arquivos grandes)
"""

import io
from openpyxl import load_workbook
from xlsxwriter import Workbook
from app.models import ProductInput, ProductOutput


def read_products(file_bytes: bytes) -> list[ProductInput]:
    """
    Lê a planilha .xlsx e extrai os produtos.
    
    Espera colunas: Descrição, EAN, NCM (case-insensitive).
    Retorna lista de ProductInput com o índice da linha.
    """
    wb = load_workbook(
        filename=io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )
    ws = wb.active

    # Mapear colunas pelo cabeçalho
    header_map: dict[str, int] = {}
    header_aliases = {
        "descricao": ["descrição", "descricao", "description", "desc", "produto", "nome"],
        "ean": ["ean", "gtin", "codigo de barras", "código de barras", "cod barras", "barcode"],
        "ncm": ["ncm", "cod ncm", "código ncm"],
    }

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for col_idx, cell_value in enumerate(first_row):
        if cell_value is None:
            continue
        normalized = str(cell_value).strip().lower().replace("_", " ")
        for field, aliases in header_aliases.items():
            if normalized in aliases:
                header_map[field] = col_idx
                break

    if "descricao" not in header_map:
        wb.close()
        raise ValueError(
            "Coluna 'Descrição' não encontrada na planilha. "
            "Colunas aceitas: " + ", ".join(header_aliases["descricao"])
        )

    # Extrair produtos
    products: list[ProductInput] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        descricao_val = row[header_map["descricao"]] if header_map.get("descricao") is not None else None

        if not descricao_val or str(descricao_val).strip() == "":
            continue

        ean_val = ""
        if "ean" in header_map and header_map["ean"] < len(row):
            ean_raw = row[header_map["ean"]]
            ean_val = str(ean_raw).strip() if ean_raw else ""
            # Remover ".0" de EANs lidos como número
            if ean_val.endswith(".0"):
                ean_val = ean_val[:-2]

        ncm_val = ""
        if "ncm" in header_map and header_map["ncm"] < len(row):
            ncm_raw = row[header_map["ncm"]]
            ncm_val = str(ncm_raw).strip() if ncm_raw else ""
            if ncm_val.endswith(".0"):
                ncm_val = ncm_val[:-2]

        products.append(ProductInput(
            row_index=row_idx,
            descricao=str(descricao_val).strip(),
            ean=ean_val,
            ncm=ncm_val,
        ))

    wb.close()
    return products


def read_feedback_products(file_bytes: bytes) -> list[dict]:
    """
    Lê a planilha de retroalimentação (corrigida manualmente).
    
    Espera colunas: Descrição, EAN, NCM, Grupo, Subgrupo.
    Retorna lista de dicts com os dados preenchidos.
    """
    wb = load_workbook(
        filename=io.BytesIO(file_bytes),
        read_only=True,
        data_only=True,
    )
    ws = wb.active

    header_map: dict[str, int] = {}
    header_aliases = {
        "descricao": ["descrição", "descricao", "description", "desc", "produto", "nome"],
        "ean": ["ean", "gtin", "codigo de barras", "código de barras"],
        "ncm": ["ncm", "cod ncm", "código ncm"],
        "grupo": ["grupo", "category", "categoria"],
        "subgrupo": ["subgrupo", "subcategoria", "subcategory", "sub grupo"],
    }

    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for col_idx, cell_value in enumerate(first_row):
        if cell_value is None:
            continue
        normalized = str(cell_value).strip().lower().replace("_", " ")
        for field, aliases in header_aliases.items():
            if normalized in aliases:
                header_map[field] = col_idx
                break

    required = ["descricao", "grupo", "subgrupo"]
    missing = [f for f in required if f not in header_map]
    if missing:
        wb.close()
        raise ValueError(f"Colunas obrigatórias não encontradas: {missing}")

    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        descricao = row[header_map["descricao"]] if header_map["descricao"] < len(row) else None
        grupo = row[header_map["grupo"]] if header_map["grupo"] < len(row) else None
        subgrupo = row[header_map["subgrupo"]] if header_map["subgrupo"] < len(row) else None

        if not descricao or not grupo or not subgrupo:
            continue

        ean_val = ""
        if "ean" in header_map and header_map["ean"] < len(row):
            ean_raw = row[header_map["ean"]]
            ean_val = str(ean_raw).strip() if ean_raw else ""
            if ean_val.endswith(".0"):
                ean_val = ean_val[:-2]

        ncm_val = ""
        if "ncm" in header_map and header_map["ncm"] < len(row):
            ncm_raw = row[header_map["ncm"]]
            ncm_val = str(ncm_raw).strip() if ncm_raw else ""
            if ncm_val.endswith(".0"):
                ncm_val = ncm_val[:-2]

        rows.append({
            "descricao": str(descricao).strip(),
            "ean": ean_val,
            "ncm": ncm_val,
            "grupo": str(grupo).strip(),
            "subgrupo": str(subgrupo).strip(),
        })

    wb.close()
    return rows


def write_results(products: list[ProductOutput]) -> io.BytesIO:
    """
    Gera o .xlsx de resultado com as colunas do input + colunas de categorização.
    Usa xlsxwriter em modo streaming (memória eficiente).
    """
    output = io.BytesIO()

    wb = Workbook(output, {"in_memory": True})
    ws = wb.add_worksheet("Resultado")

    # --- Estilos ---
    header_fmt = wb.add_format({
        "bold": True,
        "bg_color": "#1a1a2e",
        "font_color": "#e0e0ff",
        "border": 1,
        "text_wrap": True,
        "align": "center",
        "valign": "vcenter",
        "font_size": 11,
    })

    approved_fmt = wb.add_format({
        "bg_color": "#d4edda",
        "font_color": "#155724",
        "border": 1,
    })

    pending_fmt = wb.add_format({
        "bg_color": "#fff3cd",
        "font_color": "#856404",
        "border": 1,
    })

    cell_fmt = wb.add_format({
        "border": 1,
        "text_wrap": True,
        "valign": "vcenter",
    })

    # --- Cabeçalhos ---
    headers = ["Descrição", "EAN", "NCM", "Grupo", "Subgrupo", "Origem da Decisão", "Status"]
    col_widths = [50, 16, 12, 20, 25, 22, 20]

    for col, (header, width) in enumerate(zip(headers, col_widths)):
        ws.set_column(col, col, width)
        ws.write(0, col, header, header_fmt)

    # --- Dados ---
    for row_idx, product in enumerate(products, start=1):
        ws.write(row_idx, 0, product.descricao, cell_fmt)
        ws.write(row_idx, 1, product.ean, cell_fmt)
        ws.write(row_idx, 2, product.ncm, cell_fmt)
        ws.write(row_idx, 3, product.grupo, cell_fmt)
        ws.write(row_idx, 4, product.subgrupo, cell_fmt)
        ws.write(row_idx, 5, product.origem, cell_fmt)

        status_fmt = approved_fmt if product.status == "Aprovado" else pending_fmt
        ws.write(row_idx, 6, product.status, status_fmt)

    # Filtro automático
    ws.autofilter(0, 0, len(products), len(headers) - 1)

    # Congelar painel de cabeçalho
    ws.freeze_panes(1, 0)

    wb.close()
    output.seek(0)
    return output
